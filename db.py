# db.py

import sqlite3
import os
import csv
from openpyxl import Workbook, load_workbook

from config import (
    DB_PATH, CSV_PATH, XLSX_PATH,
    ADMINS
)

def init_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    cursor = conn.cursor()

    # Добавляем поле timestamp TEXT (для даты/времени транзакции).
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        worker_id INTEGER,
        withdraw_amount REAL,
        fee_amount REAL,
        owner_1_received REAL,
        owner_2_received REAL,
        status TEXT,
        screenshot_file_id TEXT,
        owner_1_withdrawn INTEGER DEFAULT 0,
        owner_2_withdrawn INTEGER DEFAULT 0,
        timestamp TEXT
    )
    """)

    # Таблица workers:
    #  - percentage (базовый процент)
    #  - owner_1_share, owner_2_share (распределение)
    #  - use_quota_logic (0/1)
    #  - daily_quota (сколько в день нужно делать, чтобы получить 20%)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS workers (
        worker_id INTEGER PRIMARY KEY,
        percentage REAL,
        owner_1_share REAL,
        owner_2_share REAL,
        use_quota_logic INTEGER DEFAULT 0,
        daily_quota REAL DEFAULT 13.0
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS admins (
        admin_id INTEGER PRIMARY KEY
    )
    """)

    # Заполним список админов из конфигурации (если есть)
    for adm in ADMINS:
        cursor.execute("INSERT OR IGNORE INTO admins (admin_id) VALUES (?)", (adm,))

    conn.commit()
    return conn


def get_worker_data(conn, worker_id):
    """
    Возвращаем:
      (percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota)
    Если нет - создаём с дефолтом: (30, 0.5, 0.5, 0, 13.0).
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota
        FROM workers
        WHERE worker_id=?
    """, (worker_id,))
    row = cursor.fetchone()
    if row:
        return row
    else:
        # Создаём
        default_p = 30.0
        default_o1 = 0.5
        default_o2 = 0.5
        default_qflag = 0
        default_dq = 13.0
        cursor.execute("""
            INSERT INTO workers (worker_id, percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota)
            VALUES (?,?,?,?,?,?)
        """, (worker_id, default_p, default_o1, default_o2, default_qflag, default_dq))
        conn.commit()
        return (default_p, default_o1, default_o2, default_qflag, default_dq)

def set_worker_percentage(conn, worker_id, percentage):
    cursor = conn.cursor()
    cursor.execute("SELECT worker_id FROM workers WHERE worker_id=?", (worker_id,))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""
            INSERT INTO workers (worker_id, percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota)
            VALUES (?,?,?,?,?,?)
        """, (worker_id, percentage, 0.5, 0.5, 0, 13.0))
    else:
        cursor.execute("""
            UPDATE workers SET percentage=? WHERE worker_id=?
        """, (percentage, worker_id))
    conn.commit()

def set_worker_owners_share(conn, worker_id, o1, o2):
    cursor = conn.cursor()
    cursor.execute("SELECT worker_id FROM workers WHERE worker_id=?", (worker_id,))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""
            INSERT INTO workers (worker_id, percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota)
            VALUES (?,?,?,?,?,?)
        """, (worker_id, 30.0, o1, o2, 0, 13.0))
    else:
        cursor.execute("""
            UPDATE workers SET owner_1_share=?, owner_2_share=? WHERE worker_id=?
        """, (o1, o2, worker_id))
    conn.commit()

def set_worker_quota_logic(conn, worker_id, enable: bool):
    val = 1 if enable else 0
    cursor = conn.cursor()
    cursor.execute("UPDATE workers SET use_quota_logic=? WHERE worker_id=?", (val, worker_id))
    conn.commit()

def set_worker_daily_quota(conn, worker_id, daily: float):
    cursor = conn.cursor()
    cursor.execute("UPDATE workers SET daily_quota=? WHERE worker_id=?", (daily, worker_id))
    conn.commit()


import datetime

def add_transaction(conn, worker_id, withdraw_amount, fee_amount,
                    owner_1_received, owner_2_received,
                    status, screenshot_file_id):
    """
    Записываем транзакцию + timestamp = now()
    owner_1_withdrawn=0, owner_2_withdrawn=0 => владелец ещё не забрал
    """
    ts = datetime.datetime.now().isoformat(timespec='seconds')
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO transactions (
            worker_id, withdraw_amount, fee_amount,
            owner_1_received, owner_2_received,
            status, screenshot_file_id,
            owner_1_withdrawn, owner_2_withdrawn,
            timestamp
        )
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (
        worker_id, withdraw_amount, fee_amount,
        owner_1_received, owner_2_received,
        status, screenshot_file_id,
        0, 0,
        ts
    ))
    conn.commit()

    # Пишем в CSV/Excel
    save_transaction_to_csv(worker_id, withdraw_amount, fee_amount, owner_1_received, owner_2_received, status, screenshot_file_id)
    save_transaction_to_excel(worker_id, withdraw_amount, fee_amount, owner_1_received, owner_2_received, status, screenshot_file_id)

def save_transaction_to_csv(worker_id, withdraw_amount, fee_amount, o1_rec, o2_rec, status, screenshot_file_id):
    file_exists = os.path.isfile(CSV_PATH)
    import csv
    with open(CSV_PATH, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow(["worker_id", "withdraw_amount", "fee_amount", "owner_1_received", "owner_2_received", "status", "screenshot_file_id"])
        writer.writerow([worker_id, withdraw_amount, fee_amount, o1_rec, o2_rec, status, screenshot_file_id])

def save_transaction_to_excel(worker_id, withdraw_amount, fee_amount, o1_rec, o2_rec, status, screenshot_file_id):
    from openpyxl import Workbook, load_workbook
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["worker_id", "withdraw_amount", "fee_amount", "owner_1_received", "owner_2_received", "status", "screenshot_file_id"])
    ws.append([worker_id, withdraw_amount, fee_amount, o1_rec, o2_rec, status, screenshot_file_id])
    wb.save(XLSX_PATH)


def get_owner_pending_sum(conn, owner_name: str) -> float:
    """
    Суммируем все транзакции, где владелец не забрал деньги.
    """
    cursor = conn.cursor()
    if owner_name == "owner1":
        cursor.execute("""
            SELECT COALESCE(SUM(owner_1_received), 0)
            FROM transactions
            WHERE owner_1_withdrawn=0
        """)
    else:
        cursor.execute("""
            SELECT COALESCE(SUM(owner_2_received), 0)
            FROM transactions
            WHERE owner_2_withdrawn=0
        """)
    row = cursor.fetchone()
    return row[0] if row else 0.0

def reset_owner_pending_sum(conn, owner_name: str):
    """
    Сбрасываем только для одного владельца.
    """
    cursor = conn.cursor()
    if owner_name == "owner1":
        cursor.execute("UPDATE transactions SET owner_1_withdrawn=1 WHERE owner_1_withdrawn=0")
    else:
        cursor.execute("UPDATE transactions SET owner_2_withdrawn=1 WHERE owner_2_withdrawn=0")
    conn.commit()


def is_admin(conn, user_id) -> bool:
    cursor = conn.cursor()
    cursor.execute("SELECT admin_id FROM admins WHERE admin_id=?", (user_id,))
    return bool(cursor.fetchone())

def get_all_workers(conn):
    """
    Выводим всех работников (если нужно).
    worker_id, percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT worker_id, percentage, owner_1_share, owner_2_share, use_quota_logic, daily_quota
        FROM workers
    """)
    return cursor.fetchall()


def get_last_transaction_timestamp(conn, worker_id: int):
    """
    Возвращает timestamp (TEXT) последней транзакции для данного worker_id,
    в формате ISO 'YYYY-MM-DDTHH:MM:SS', или None, если нет.
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT timestamp
        FROM transactions
        WHERE worker_id=?
        ORDER BY id DESC LIMIT 1
    """, (worker_id,))
    row = cursor.fetchone()
    if row:
        return row[0]  # строка
    else:
        return None
